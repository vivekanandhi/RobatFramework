import openpyxl


vk = openpyxl.load_workbook("C:/Users/46734/Documents/vivi/userpass100.xlsx")


def fetch_number_of_rows(sheetname):
    sh = vk[sheetname]
    return sh.max_row


def fetch_cell_data(sheetname,row,cell):
    sh = vk[sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value






